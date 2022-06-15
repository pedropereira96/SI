from itertools import count
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
import time
import random
from selenium.webdriver.common.by import By
from selenium.webdriver.firefox.service import Service

from openpyxl import Workbook, load_workbook

#countries = ["Argentina","Australia","Brazil","Canada","China","Egypt","France","Germany","India","Italy","Japan","Mexico","Portugal","South Africa","Spain","United Kingdom","United States"]
capitals = ["Buenos Aires", "Camberra", "Brasilia", "Ottawa", "Beijing", "Cairo", "Paris", "Berlin", "New Delhi", "Rome", "Tokyo", "Mexico City", "Lisbon", "Cape Town", "Madrid", "London", "Washington D.C."]
continents = ["America", "Australia", "America", "America", "Asia", "Africa", "Europe", "Europe", "Asia", "Europe", "Asia", "America", "Europe", "Africa", "Europe", "Europe", "America"]
seasons = ["Summer", "Winter", "Autumn", "Spring"]

wb = Workbook()


#Flights

classes = ["Economy", "Executive", "First Class"]
jun_days = ["20-06", "22-06","23-06", "25-06", "26-06", "28-06", "30-06"]
jul_days = ["01-07", "02-07", "03-07", "04-07", "06-07", "07-07", "09-07", "10-07", "12-07", "14-07", "16-07", "17-07", "18-07", "19-07", "21-07", "22-07", "24-07", "26-07", "28-07", "30-07", "31-07"]
hours = ["6:30", "13:00", "17:00"]

ws1 = wb.active
ws1.title = "Flights"
header = ("From - Country", "From - Continent", "To - Country", "To - Continent", "Date", "Hour", "Class", "Price")
ws1.append(header)

def write_flights(country1, continent1, country2, continent2, month):
    for d in month:
        for h in hours:
            price = random.uniform(100,500)
            for c in classes:
                new_line = (country1, continent1, country2, continent2, d, h, c, price)
                ws1.append(new_line)
                price = price + random.uniform(50,500)   

for c1 in range(len(capitals)):
    for c2 in range(len(capitals)):
        if (c1 != c2):
            country1 = capitals[c1]
            country2 = capitals[c2]
            continent1 = continents[c1]
            continent2 = continents[c2]
             
            write_flights(country1, continent1, country2, continent2, jun_days)
            write_flights(country1, continent1, country2, continent2, jul_days)

###############################################################################################################################

# Seasons
ws2 = wb.create_sheet(title='Seasons')
header = ("To", "Season", "date_start", "date_end")
ws2.append(header)

for c in capitals:
    for s in seasons:
        new_line = (c, s)
        ws2.append(new_line)

###############################################################################################################################

# Accomondations

ws3 = wb.create_sheet(title='Accomondations')
header = ("Where", "Season", "Type", "Price")
ws3.append(header)

accomodation_type = ["Camping", "Hostel","House","Hotel"]

for country in capitals:
    for s in seasons:
        price = random.uniform(10,60)
        for c in accomodation_type:
            
            new_line = (country, s, c, price)
            ws3.append(new_line)
            price = price + random.uniform(50,200)

###############################################################################################################################

# Attractions

ws4 = wb.create_sheet(title='Attractions')
header = ("Where", "Season", "Type", "Name", "Price")
ws4.append(header)

###############################################################################################################################

wb.save(filename='dataset_v3.xlsx') 
